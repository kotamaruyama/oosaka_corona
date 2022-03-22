#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import glob
import re
import openpyxl
import pandas as pd
import numpy as np


# In[2]:


pwd


# In[3]:


cd Downloads/


# In[25]:


filepath_list = glob.glob(os.getcwd() + '/oosaka/*.xlsx')
filepath_list


# In[26]:


filename_list = []
dir_path = ''
for i, mypath in enumerate(filepath_list):
    mypath = str(mypath)
    mydir_path, file_name = os.path.split(mypath)
    filename_list.append(file_name)
    if i == 0:
        dir_path = mydir_path


# In[27]:


filename_list


# In[28]:


filename_list_sorted = sorted(filename_list, key=lambda x:int((re.search(r"[0-9]+", x)).group(0)))
print(filename_list_sorted)


# In[29]:


new_filepath_list = []
for myfilename in filename_list_sorted:
    mypath = os.path.join(dir_path, myfilename)
    new_filepath_list.append(mypath)
new_filepath_list


# In[30]:


new_filepath_list.remove("/Users/maruyamakouta/Downloads/oosaka/~$0319.xlsx")


# In[31]:


new_filepath_list


# In[54]:


# エクセルのセル行列範囲の表データを抽出する
df_list = []
for k, filepath in enumerate(new_filepath_list, start=1):
    # エクセルファイルのインスタンス生成（オブジェクト）
    # 引数data_only=Trueで、セル内が数式の場合であっても数値で取得する
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # ワークシートを読み込む
    ws = wb.worksheets[1]
    # ws = wb.get_sheet_by_name('シート名')
    
    # 行番号を検索により取得する
    start_row = ''
    
    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        # A列（pythonでは0番目）を対象に、キーワード検索して、その行番号を取得する
        if row[0].value == '大阪市': # A列で上から検索して、最小にxを見つけた時
            start_row = i
            
            break # forループを抜ける

    # エクセルシート内にある最終行を取得する
    end_row = ws.max_row
    #print('start_row, end_row', start_row, end_row)
    
    # 指定行の最終列の列番号を取得
    end_col = ''
    for j in reversed(range(1, ws.max_column)):
        if ws.cell(row=start_row, column=j).value != None:
            end_col =j
            #print('end_col', end_col)
            break
    
    # 表範囲を配列で取得する
    data_rows = []
    for rows in ws.iter_rows(min_row = start_row, min_col = 1,
                             max_row = end_row, max_col = end_col+1):
        data_cols = []
        for cell in rows:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    
    print(data_rows)
    # pandasデータフレーム形式へ変換
    df = pd.DataFrame(data_rows)
    #print(df)
    
    # 1行目を取得
    #column_names = df.iloc[0]
    
    # 列名を置換する。そして、元の1行目は削除
   # df.columns = column_names
    
    #print(df2)
    
    # カウント数をカテゴリ変数として入れる
   # my_label = 'category'
    #df2[my_label] = k
    #print(df2)
    
    # dfをリストへ格納
    #df_list.append(df2)
   


# In[55]:


df.head(100)


# In[46]:


df.to_csv("oosaka24.csv", encoding="shift_jis")


# In[ ]:




